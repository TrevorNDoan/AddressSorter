import csv
import openpyxl
import os
import requests
from PIL import Image
from io import BytesIO

# Google API Requirements
API_KEY = ""
size = "640x640"

# Open the CSV file and create the sorted Excel spreadsheet
csv_file = open('AddressList.csv', 'r', encoding='utf-8-sig')
if os.path.exists('sorted.xlsx'):
    workbook = openpyxl.load_workbook('sorted.xlsx')
else:
    workbook = openpyxl.Workbook()
    house_sheet = workbook.active
    house_sheet.title = 'House'
    office_sheet = workbook.create_sheet('Office')
    misc_sheet = workbook.create_sheet('Misc')
    sorted_sheet = workbook.create_sheet(title='Sorted')
workbook.save('sorted.xlsx')

# Get references to the existing sheets
house_sheet = workbook['House']
office_sheet = workbook['Office']
misc_sheet = workbook['Misc']
sorted_sheet = workbook['Sorted']

# Loop through each row in the CSV file
csv_reader = csv.DictReader(csv_file)
rows = list(csv_reader)  # Convert the CSV reader to a list of rows for indexing
for i, row in enumerate(rows):
    full_address = row['fullAddress']
    print(full_address)
    
    url = f"https://maps.googleapis.com/maps/api/streetview?size={size}&location={full_address}&key={API_KEY}"
    
    response = requests.get(url)
    img = Image.open(BytesIO(response.content))
    
    img.show()

    # Prompt the user for the building type
    while True:
        building_type = input('Is this a House, Office or Misc? (Enter "Exit" to quit) ')
        if building_type.lower() in ['house', 'office', 'misc']:
            # Convert the row values to a list and add the row to the appropriate sheet in the Excel spreadsheet
            row_list = list(row.values())
            if building_type.lower() == 'house':
                house_sheet.append(row_list)
            elif building_type.lower() == 'office':
                office_sheet.append(row_list)
            elif building_type.lower() == 'misc':
                misc_sheet.append(row_list)
            elif building_type.lower() == 'sorted':
                sorted_sheet.append(row_list)
            # Remove the selected row from the CSV file
            rows.pop(i)
            break
        elif building_type.lower() == 'exit':
            # Save the Excel spreadsheet and exit
            workbook.save('sorted.xlsx')
            # Write the updated CSV data to a new file
            with open('UpdatedAddressTypes.csv', 'w', newline='', encoding='utf-8-sig') as csv_out:
                writer = csv.DictWriter(csv_out, fieldnames=csv_reader.fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            # Close the CSV file and delete it
            csv_file.close()
            os.remove('AddressList.csv')
            # Rename the updated CSV file to the original name
            os.rename('UpdatedAddressTypes.csv', 'AddressList.csv')
            exit()
        else:
            print('Invalid input. Please enter "House", "Office", "Misc", "Sorted", or "Exit".')

# Save the Excel spreadsheet
workbook.save('sorted.xlsx')

# Save the Excel spreadsheet
workbook.save('sorted.xlsx')

# Write the updated CSV data to a new file
with open('UpdatedAddressTypes.csv', 'w', newline='', encoding='utf-8-sig') as csv_out:
    writer = csv.DictWriter(csv_out, fieldnames=csv_reader.fieldnames)
    writer.writeheader()
    writer.writerows(rows)

# Delete the old CSV file and rename the updated file to the original name
os.remove('AddressList.csv')
os.rename('UpdatedAddressTypes.csv', 'AddressList.csv')